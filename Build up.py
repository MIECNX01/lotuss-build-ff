import openpyxl
import os
import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import pythoncom
import win32com.client as win32
from collections import Counter
from openpyxl.drawing.image import Image  # สำหรับแทรกรูปโลโก้

# -----------------------------
# Path
current_dir = os.path.dirname(os.path.abspath(__file__))
template_path = os.path.join(current_dir, "Asset", "ไฟล์ต้นฉบับ.xlsx")
item_file_path = os.path.join(current_dir, "Asset", "ข้อมูลไอเท็ม.xlsx")
logo_path = os.path.join(current_dir, "Asset", "logo.jpg")
output_dir = os.path.join(current_dir, "Out Build")
start_row = 17
# -----------------------------
os.makedirs(output_dir, exist_ok=True)

# โหลดข้อมูลไอเท็มจากคอลัมน์ M -> N และ M -> E (เพื่อดูชื่อแผนกด้วย)
item_dict = {}        # รหัส → ชื่อสินค้า
item_dept_dict = {}   # รหัส → แผนก
if os.path.exists(item_file_path):
    wb_items = openpyxl.load_workbook(item_file_path, data_only=True)
    ws_items = wb_items.active
    for row in ws_items.iter_rows(min_row=2, values_only=True):
        item_code = str(row[12]).strip() if row[12] else ""
        item_name = str(row[13]).strip() if row[13] else ""
        item_dept = str(row[4]).strip() if row[4] else ""
        if item_code:
            if item_name:
                item_dict[item_code] = item_name
            if item_dept:
                item_dept_dict[item_code] = item_dept

code_entries = []
name_entries = []
avg_entries = []
qty_entries = []
price_entries = []

# -----------------------------
def search_item_name(idx):
    code = code_entries[idx].get().strip()
    name = item_dict.get(code, "")
    name_entries[idx].delete(0, tk.END)
    name_entries[idx].insert(0, name)

def add_row():
    row = len(code_entries)
    y = 7 + row

    code_e = tk.Entry(root, width=15)
    name_e = tk.Entry(root, width=25)
    avg_e = tk.Entry(root, width=10)
    qty_e = tk.Entry(root, width=10)
    price_e = tk.Entry(root, width=10)
    search_btn = tk.Button(root, text="ค้นหาชื่อ", command=lambda idx=row: search_item_name(idx))

    code_e.grid(row=y, column=0)
    name_e.grid(row=y, column=1)
    avg_e.grid(row=y, column=2)
    qty_e.grid(row=y, column=3)
    price_e.grid(row=y, column=4)
    search_btn.grid(row=y, column=5)

    code_entries.append(code_e)
    name_entries.append(name_e)
    avg_entries.append(avg_e)
    qty_entries.append(qty_e)
    price_entries.append(price_e)

def save_to_excel():
    try:
        branch_code = branch_entry.get().strip()
        branch_name = branchname_entry.get().strip()

        if not branch_code.isdigit() or len(branch_code) != 4:
            messagebox.showerror("❌ ผิดพลาด", "กรุณากรอกรหัสสาขา 4 หลักให้ถูกต้อง")
            return
        if not branch_name:
            messagebox.showerror("❌ ผิดพลาด", "กรุณากรอกชื่อสาขา")
            return

        dept_list = []
        for i in range(len(code_entries)):
            code = code_entries[i].get().strip()
            if code in item_dept_dict:
                dept_list.append(item_dept_dict[code])

        if not dept_list:
            messagebox.showerror("❌ ผิดพลาด", "ไม่พบแผนกสินค้าใดเลยจากรหัสที่กรอก")
            return

        dept_count = Counter(dept_list)
        most_common = dept_count.most_common()
        sheet_name = most_common[0][0]

        if len(most_common) > 1 and most_common[0][1] == most_common[1][1]:
            first_code = code_entries[0].get().strip()
            sheet_name = item_dept_dict.get(first_code, sheet_name)

        dept_label_var.set(f"ชื่อแผนก: {sheet_name}")

        wb = openpyxl.load_workbook(template_path)
        if sheet_name not in wb.sheetnames:
            template_sheet = None
            for sht in wb.sheetnames:
                if 'เปลี่ยน' in sht or 'template' in sht.lower():
                    template_sheet = wb[sht]
                    break
            if not template_sheet:
                template_sheet = wb.worksheets[0]
            new_sheet = wb.copy_worksheet(template_sheet)
            new_sheet.title = sheet_name

        # ลบชีตอื่นออก เหลือแค่แผนกเดียว
        for sheet in wb.sheetnames:
            if sheet != sheet_name:
                wb.remove(wb[sheet])

        today_excel = datetime.now().strftime("%d/%m/%y")
        today_filename = datetime.now().strftime("%d-%m-%Y")
        ws = wb[sheet_name]
        ws['C9'] = branch_code
        ws['C10'] = branch_name
        ws['C11'] = today_excel
        ws['C12'] = sheet_name

        # แทรกรูปโลโก้ลง A1
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.width = 180
            img.height = 55
            ws.add_image(img, "A1")

        # ใส่คำว่า Memo Randum ลงในเซลล์แบบปรับขนาด
        ws.merge_cells('F1:I2')
        ws['F1'] = "Memo Randum"
        ws['F1'].font = openpyxl.styles.Font(size=28, bold=True)
        ws['F1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

        row = start_row
        while ws.cell(row=row, column=1).value:
            row += 1

        added = 0
        for i in range(len(code_entries)):
            code = code_entries[i].get().strip()
            name = name_entries[i].get().strip()
            avg = avg_entries[i].get().strip()
            qty = qty_entries[i].get().strip()
            price = price_entries[i].get().strip()

            if code and name and avg and qty and price:
                avg_f = float(avg)
                qty_f = int(qty)
                price_f = float(price)

                ws.cell(row=row + added, column=1).value = branch_code
                ws.cell(row=row + added, column=2).value = added + 1
                ws.cell(row=row + added, column=3).value = code
                ws.cell(row=row + added, column=4).value = name
                ws.cell(row=row + added, column=5).value = 0
                ws.cell(row=row + added, column=6).value = today_excel
                ws.cell(row=row + added, column=7).value = qty_f
                ws.cell(row=row + added, column=8).value = 0
                ws.cell(row=row + added, column=9).value = "สินค้าไม่เพียงพอต่อการขาย"
                ws.cell(row=row + added, column=10).value = avg_f
                ws.cell(row=row + added, column=11).value = round(avg_f / 7, 2)
                ws.cell(row=row + added, column=12).value = 0
                ws.cell(row=row + added, column=13).value = price_f
                ws.cell(row=row + added, column=14).value = round(qty_f * price_f, 2)
                added += 1

        if added == 0:
            messagebox.showerror("❌ ผิดพลาด", "กรุณากรอกข้อมูลสินค้าอย่างน้อย 1 รายการ")
            return

        for r in range(row + added, row + 100):
            ws.row_dimensions[r].hidden = True

        # สร้าง subfolder ตามวัน เวลา ชื่อแผนก และลำดับครั้ง
        from glob import glob
        subfolder_base = os.path.join(output_dir, f"{today_filename}_{sheet_name}")
        os.makedirs(subfolder_base, exist_ok=True)
        existing = len(glob(os.path.join(subfolder_base, "*.xlsx")))
        run_number = existing + 1
        subfolder_path = os.path.join(subfolder_base, str(run_number))
        os.makedirs(subfolder_path, exist_ok=True)

        export_name = f"แบบฟอร์มเพิ่มออร์เดอร์ FF {today_filename} {sheet_name} #{branch_code}"
        xlsx_path = os.path.join(subfolder_path, export_name + ".xlsx")
        pdf_path = os.path.join(subfolder_path, export_name + ".pdf")

        wb.save(xlsx_path)

        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb_excel = excel.Workbooks.Open(xlsx_path)
        excel.DisplayAlerts = False

        for sht in wb_excel.Sheets:
            if sht.Name == sheet_name:
                sht.PageSetup.Orientation = 2
                sht.PageSetup.Zoom = False
                sht.PageSetup.FitToPagesWide = 1
                sht.PageSetup.FitToPagesTall = False
                sht.ExportAsFixedFormat(0, pdf_path)

        wb_excel.Close(False)
        excel.Quit()

        messagebox.showinfo("✅ สำเร็จ", f"บันทึกแล้ว:\n{xlsx_path}\n{pdf_path}")
        root.destroy()

    except Exception as e:
        messagebox.showerror("❌ ผิดพลาด", str(e))


# -----------------------------
def delete_last_row():
    if code_entries:
        code_entries.pop().destroy()
        name_entries.pop().destroy()
        avg_entries.pop().destroy()
        qty_entries.pop().destroy()
        price_entries.pop().destroy()
        # Remove the last search button
        for widget in root.grid_slaves():
            if int(widget.grid_info()["row"]) == 6 + len(code_entries):
                if isinstance(widget, tk.Button) and widget.cget("text") == "ค้นหาชื่อ":
                    widget.destroy()
                    break

# -----------------------------
# GUI
root = tk.Tk()
root.title("กรอกสินค้าเข้าแผนก")

tk.Label(root, text="รหัสสาขา (4 หลัก)").grid(row=0, column=0)
tk.Label(root, text="ชื่อสาขา").grid(row=1, column=0)

dept_label_var = tk.StringVar()
dept_label = tk.Label(root, textvariable=dept_label_var, fg="blue")
dept_label.grid(row=2, column=0, columnspan=6, pady=(0, 5))

branch_entry = tk.Entry(root)
branchname_entry = tk.Entry(root)
branch_entry.grid(row=0, column=1, columnspan=4, sticky="we", pady=2)
branchname_entry.grid(row=1, column=1, columnspan=4, sticky="we", pady=2)

headers = ["รหัสสินค้า", "ชื่อสินค้า", "ยอดขาย/สัปดาห์", "จำนวนเพิ่ม", "ราคาต่อหน่วย", ""]
for col, h in enumerate(headers):
    tk.Label(root, text=h).grid(row=6, column=col)

add_row()

tk.Button(root, text="➕ เพิ่มแถว", command=add_row).grid(row=999, column=0, pady=10)
tk.Button(root, text="❌ ลบแถว", command=lambda: delete_last_row()).grid(row=999, column=1, pady=10)
tk.Button(root, text="📤 สรุปส่งออก", command=save_to_excel).grid(row=999, column=2, columnspan=3, pady=10)

root.mainloop()
